from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data" / "operations"
TARGET_XLSX = Path(r"C:\Users\rodolfo.crotti\AXS ENERGIA S A\Financeiro - Documentos\000 - FD\Rodolfo relatórios\PMT.xlsx")
CHART_SHEET_NAME = "Graficos PPT"
CHART_DATA_SHEET_NAME = "_Base Graficos"
EXPORT_FILES = [
    ("axs01--primeira.json", "1ª Emissão"),
    ("axs01--segunda.json", "2ª Emissão"),
    ("axs02--cri.json", "CRI"),
    ("axs02--deb.json", "Debênture"),
    ("axs03.json", ""),
    ("axs04.json", ""),
    ("axs05--primeira.json", "1ª Emissão"),
    ("axs05--segunda.json", "2ª Emissão"),
    ("axs06--primeira.json", "1ª Emissão"),
    ("axs06--segunda.json", "2ª Emissão"),
    ("axs07.json", ""),
    ("axs08.json", ""),
    ("axs09.json", ""),
    ("axs10.json", ""),
    ("axs11.json", ""),
    ("axsgoias.json", ""),
]
COMPANY_REGISTRY = {
    "axs01": {"number": 10, "name": "AXS ENERGIA UNIDADE 01 LTDA"},
    "axs02": {"number": 11, "name": "AXS ENERGIA UNIDADE 02 S.A."},
    "axs03": {"number": 16, "name": "AXS ENERGIA UNIDADE 03 LTDA"},
    "axs04": {"number": 17, "name": "AXS ENERGIA UNIDADE 04 S.A."},
    "axs05": {"number": 18, "name": "AXS ENERGIA UNIDADE 05 S.A."},
    "axs06": {"number": 19, "name": "AXS ENERGIA UNIDADE 06 S.A."},
    "axs07": {"number": 25, "name": "AXS ENERGIA UNIDADE 07 S.A."},
    "axs08": {"number": 26, "name": "AXS ENERGIA UNIDADE 08 S.A."},
    "axs09": {"number": 27, "name": "AXS ENERGIA UNIDADE 09 S.A."},
    "axs10": {"number": 28, "name": "AXS ENERGIA UNIDADE 10 S.A."},
    "axs11": {"number": 29, "name": "AXS ENERGIA UNIDADE 11 S.A."},
    "axsgoias": {"number": 32, "name": "AXS ENERGIA UFV GOIAS SPE LTDA"},
}
HEADERS = [
    "Número empresa",
    "Empresa",
    "Data de pagamento",
    "Operação",
    "Emissão / instrumento",
    "Valor PMT",
    "Juros",
    "Amortização",
    "Saldo após pagamento",
    "Indexador",
    "Atualizado em",
]


def parse_br_date(value: str) -> datetime:
    return datetime.strptime(value, "%d/%m/%Y")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def company_info(operation_id: str, fallback_name: str) -> dict[str, str | int]:
    info = COMPANY_REGISTRY.get(operation_id, {})
    return {
        "number": info.get("number", ""),
        "name": info.get("name", fallback_name),
    }


def future_payment_rows(payload: dict, emission_label: str) -> list[dict]:
    today = datetime.now().date()
    operation = payload["operation"]
    operation_id = operation.get("id") or ""
    company = company_info(operation_id, operation.get("issuer") or operation.get("label") or "-")
    rows: list[dict] = []
    for item in payload.get("table_series", []):
        date_text = item.get("date")
        payment = item.get("payment") or 0
        if not date_text or payment <= 0:
            continue
        payment_date = parse_br_date(date_text).date()
        if payment_date < today:
            continue
        rows.append(
            {
                "date": payment_date,
                "date_text": date_text,
                "company_number": company["number"],
                "company_name": company["name"],
                "operation": operation.get("label") or "-",
                "emission": emission_label or item.get("component_label") or operation.get("badge") or "-",
                "payment": payment,
                "interest": item.get("interest") or 0,
                "amortization": item.get("amortization") or 0,
                "balance": item.get("balance") or 0,
                "indexer": operation.get("indexer") or "-",
            }
        )
    return rows


def collect_rows() -> list[dict]:
    rows: list[dict] = []
    for filename, emission_label in EXPORT_FILES:
        path = DATA_DIR / filename
        if not path.exists():
            continue
        payload = load_json(path)
        rows.extend(future_payment_rows(payload, emission_label))
    rows.sort(
        key=lambda item: (
            item["date"],
            str(item["company_number"]),
            str(item["company_name"]),
            item["operation"],
            item["emission"],
        )
    )
    return rows


def to_millions(value: float | int | None) -> float:
    return round((value or 0) / 1_000_000, 4)


def month_label(value) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    return str(value)


def parse_series_date(item: dict):
    value = item.get("date")
    if not value:
        return None
    return parse_br_date(value).date()


def focus_slice(series: list[dict], limit: int = 24) -> list[dict]:
    today = datetime.now().date()
    dated = [item for item in series if parse_series_date(item)]
    if not dated:
        return series[:limit]
    first_future = next((index for index, item in enumerate(dated) if parse_series_date(item) >= today), None)
    if first_future is None:
        return dated[-limit:]
    start = max(0, first_future - 2)
    return dated[start:start + limit]


def load_portfolio_payload() -> dict:
    path = DATA_DIR / "geral.json"
    return load_json(path) if path.exists() else {}


def load_operation_payloads() -> list[dict]:
    payloads = []
    for filename, _ in EXPORT_FILES:
        path = DATA_DIR / filename
        if path.exists():
            payloads.append(load_json(path))
    return payloads


def build_chart_data(rows: list[dict]) -> dict[str, list[list]]:
    portfolio = load_portfolio_payload()
    payloads = load_operation_payloads()
    today = datetime.now().date()
    series = focus_slice(portfolio.get("series", []), 16)

    event_points = []
    for item in series:
        event_points.append([
            item.get("date", "-"),
            to_millions(item.get("payment")),
            to_millions(item.get("interest")),
            to_millions(item.get("amortization")),
            to_millions(item.get("balance")),
        ])
    pmt_values = [point[1] for point in event_points if point[1]]
    balance_values = [point[4] for point in event_points if point[4]]
    pmt_average = round(sum(pmt_values) / len(pmt_values), 4) if pmt_values else 0
    balance_average = round(sum(balance_values) / len(balance_values), 4) if balance_values else 0
    event_rows = [["Data", "PMT", "Juros", "Amortização", "Saldo", "PMT médio", "Saldo médio"]]
    for point in event_points:
        event_rows.append(point + [pmt_average, balance_average])
    event_pmt_rows = [["Data", "PMT", "PMT médio"]]
    event_balance_rows = [["Data", "Saldo", "Saldo médio"]]
    for point in event_points:
        event_pmt_rows.append([point[0], point[1], pmt_average])
        event_balance_rows.append([point[0], point[4], balance_average])

    comparison = portfolio.get("comparison", [])
    if not comparison:
        comparison = [
            {
                "label": payload.get("operation", {}).get("label", "-"),
                "current_balance": payload.get("summary", {}).get("current_balance") or 0,
                "indexer": payload.get("operation", {}).get("indexer", "-"),
            }
            for payload in payloads
        ]
    comparison = sorted(comparison, key=lambda item: item.get("current_balance") or 0, reverse=True)
    balance_rows = [["Operação", "Saldo atual"]]
    for item in comparison[:12]:
        balance_rows.append([item.get("label", "-"), to_millions(item.get("current_balance"))])

    pmt_12m_by_operation: dict[str, float] = defaultdict(float)
    monthly_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"PMT": 0.0, "Juros": 0.0, "Amortização": 0.0})
    cutoff = today.replace(year=today.year + 1)
    for item in rows:
        payment_date = item["date"]
        if payment_date > cutoff:
            continue
        operation = item.get("operation") or "-"
        pmt_12m_by_operation[operation] += item.get("payment") or 0
        month = month_label(payment_date)
        monthly_totals[month]["PMT"] += item.get("payment") or 0
        monthly_totals[month]["Juros"] += item.get("interest") or 0
        monthly_totals[month]["Amortização"] += item.get("amortization") or 0

    pmt_operation_rows = [["Operação", "PMT 12 meses"]]
    for operation, amount in sorted(pmt_12m_by_operation.items(), key=lambda pair: pair[1], reverse=True)[:12]:
        pmt_operation_rows.append([operation, to_millions(amount)])

    monthly_rows = [["Mês", "PMT", "Juros", "Amortização"]]
    for month, amounts in sorted(monthly_totals.items())[:12]:
        monthly_rows.append([month, to_millions(amounts["PMT"]), to_millions(amounts["Juros"]), to_millions(amounts["Amortização"])])

    indexer_totals: dict[str, float] = defaultdict(float)
    for item in comparison:
        indexer = item.get("indexer") or "-"
        indexer_totals[indexer] += item.get("current_balance") or 0
    indexer_rows = [["Indexador", "Saldo atual"]]
    for indexer, amount in sorted(indexer_totals.items(), key=lambda pair: pair[1], reverse=True):
        indexer_rows.append([indexer, to_millions(amount)])

    return_rows = [["Operação", "TIR", "Taxa contratada", "Spread"]]
    for payload in payloads:
        summary = payload.get("summary", {})
        operation = payload.get("operation", {}).get("label", "-")
        tir = summary.get("tir_annual_pct")
        contracted = summary.get("contracted_rate_annual_pct")
        spread = summary.get("effective_spread_annual_pct")
        if tir is None:
            continue
        return_rows.append([
            operation,
            (tir or 0) / 100,
            (contracted or 0) / 100 if contracted is not None else None,
            (spread or 0) / 100 if spread is not None else None,
        ])

    summary = portfolio.get("summary", {})
    kpis = [
        ["Indicador", "Valor"],
        ["Saldo atual carteira", to_millions(summary.get("current_balance"))],
        ["PMT próximos 12 meses", to_millions(sum(item.get("payment") or 0 for item in rows if item["date"] <= cutoff))],
        ["Juros próximos 12 meses", to_millions(sum(item.get("interest") or 0 for item in rows if item["date"] <= cutoff))],
        ["Amortização próximos 12 meses", to_millions(sum(item.get("amortization") or 0 for item in rows if item["date"] <= cutoff))],
        ["TIR carteira", (summary.get("tir_annual_pct") or 0) / 100 if summary.get("tir_annual_pct") is not None else None],
        ["Spread efetivo carteira", (summary.get("effective_spread_annual_pct") or 0) / 100 if summary.get("effective_spread_annual_pct") is not None else None],
        ["Atualizado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
    ]

    return {
        "event_rows": event_rows,
        "event_pmt_rows": event_pmt_rows,
        "event_balance_rows": event_balance_rows,
        "balance_rows": balance_rows,
        "pmt_operation_rows": pmt_operation_rows,
        "monthly_rows": monthly_rows,
        "indexer_rows": indexer_rows,
        "return_rows": return_rows,
        "kpis": kpis,
    }


def write_matrix(sheet, start_row: int, start_col: int, values: list[list]) -> tuple[int, int, int, int]:
    for row_offset, row in enumerate(values):
        for col_offset, value in enumerate(row):
            sheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
    end_row = start_row + len(values) - 1
    end_col = start_col + len(values[0]) - 1 if values else start_col
    return start_row, start_col, end_row, end_col


def style_table(sheet, bounds: tuple[int, int, int, int], table_name: str | None = None) -> None:
    start_row, start_col, end_row, end_col = bounds
    header_fill = PatternFill(fill_type="solid", fgColor="12324D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for cell in sheet.iter_rows(min_row=start_row, max_row=start_row, min_col=start_col, max_col=end_col).__next__():
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if table_name and end_row > start_row:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)


def clear_generated_sheets(workbook: Workbook) -> None:
    for name in (CHART_SHEET_NAME, CHART_DATA_SHEET_NAME):
        if name in workbook.sheetnames:
            del workbook[name]


def setup_dashboard_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = None
    sheet.sheet_view.zoomScale = 80
    sheet.print_area = "A1:T76"
    if sheet.sheet_properties.pageSetUpPr is None:
        sheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 2
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    for col in range(1, 21):
        sheet.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 96):
        sheet.row_dimensions[row].height = 21
    sheet.merge_cells("A1:T2")
    sheet["A1"] = "Painel executivo de PMTs e dívida | visão para PPT"
    sheet["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="0B1F33")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("A3:T3")
    sheet["A3"] = "Fonte: mesmos dados gerados diariamente para o dashboard web e para a PMT.xlsx. Valores em R$ milhões, exceto taxas."
    sheet["A3"].font = Font(size=10, color="607D95")


def write_kpi_cards(sheet, chart_data: dict[str, list[list]]) -> None:
    kpi_map = {row[0]: row[1] for row in chart_data["kpis"][1:]}
    cards = [
        ("Saldo atual carteira", kpi_map.get("Saldo atual carteira"), "R$ mi"),
        ("PMT próximos 12 meses", kpi_map.get("PMT próximos 12 meses"), "R$ mi"),
        ("Juros próximos 12 meses", kpi_map.get("Juros próximos 12 meses"), "R$ mi"),
        ("Amortização próximos 12 meses", kpi_map.get("Amortização próximos 12 meses"), "R$ mi"),
        ("TIR carteira", kpi_map.get("TIR carteira"), "% a.a."),
        ("Spread efetivo carteira", kpi_map.get("Spread efetivo carteira"), "% a.a."),
    ]
    anchors = ["A5", "D5", "G5", "J5", "M5", "P5"]
    for anchor, (label, value, unit) in zip(anchors, cards):
        cell = sheet[anchor]
        row = cell.row
        col = cell.column
        sheet.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 1)
        target = sheet.cell(row=row, column=col)
        if value is None:
            display = "-"
        elif unit.startswith("%"):
            display = f"{value:.2%}"
        else:
            display = f"{value:,.1f}"
        target.value = f"{label}\n{display} {unit}"
        target.fill = PatternFill(fill_type="solid", fgColor="EAF3F8")
        target.font = Font(color="0B1F33", bold=True, size=11)
        target.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        target.border = Border(
            left=Side(style="thin", color="BFD4E5"),
            right=Side(style="thin", color="BFD4E5"),
            top=Side(style="thin", color="BFD4E5"),
            bottom=Side(style="thin", color="BFD4E5"),
        )


def make_gridlines() -> ChartLines:
    gridlines = ChartLines()
    gridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill="D9E5EF", w=7000)
    )
    return gridlines


def set_chart_common(chart, title: str, width: float = 15.6, height: float = 8.4) -> None:
    chart.title = title
    chart.width = width
    chart.height = height
    chart.legend.position = "b"
    chart.style = 10
    chart.y_axis.majorGridlines = make_gridlines()
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.majorTickMark = "none"


def color_series(chart, colors: list[str]) -> None:
    for series, color in zip(chart.series, colors):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 25000
        if hasattr(series, "marker"):
            series.marker.symbol = "circle"
            series.marker.size = 6


def style_reference_series(series, color: str = "7A869A") -> None:
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.dashStyle = "dash"
    series.graphicalProperties.line.width = 18000
    if hasattr(series, "marker"):
        series.marker.symbol = "none"


def add_data_labels(chart, number_format: str, position: str = "bestFit", show_series_name: bool = False) -> None:
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showSerName = show_series_name
    chart.dLbls.showLegendKey = False
    chart.dLbls.numFmt = number_format
    chart.dLbls.position = position


def add_line_chart(sheet, data_sheet, bounds, anchor: str, title: str, value_col: int, color: str, number_format: str = "#,##0.0", level_col: int | None = None, show_labels: bool = False) -> None:
    start_row, start_col, end_row, _ = bounds
    chart = LineChart()
    if level_col:
        data = Reference(data_sheet, min_col=start_col + value_col - 1, max_col=start_col + level_col - 1, min_row=start_row, max_row=end_row)
    else:
        data = Reference(data_sheet, min_col=start_col + value_col - 1, min_row=start_row, max_row=end_row)
    cats = Reference(data_sheet, min_col=start_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    set_chart_common(chart, title)
    chart.y_axis.numFmt = number_format
    chart.x_axis.tickLblSkip = 3
    chart.x_axis.tickMarkSkip = 3
    color_series(chart, [color, "7A869A"] if level_col else [color])
    if level_col and len(chart.series) > 1:
        style_reference_series(chart.series[-1])
    if show_labels:
        add_data_labels(chart, number_format, position="t")
    sheet.add_chart(chart, anchor)


def add_bar_chart(sheet, data_sheet, bounds, anchor: str, title: str, min_col_offset: int, max_col_offset: int, colors: list[str], chart_type: str = "col", number_format: str = "#,##0.0", show_legend: bool = True, show_labels: bool = True) -> None:
    start_row, start_col, end_row, _ = bounds
    chart = BarChart()
    chart.type = chart_type
    data = Reference(data_sheet, min_col=start_col + min_col_offset - 1, max_col=start_col + max_col_offset - 1, min_row=start_row, max_row=end_row)
    cats = Reference(data_sheet, min_col=start_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    set_chart_common(chart, title)
    chart.overlap = 0
    chart.gapWidth = 65 if chart_type == "bar" else 95
    chart.y_axis.numFmt = number_format
    if chart_type == "bar":
        chart.x_axis.numFmt = number_format
    color_series(chart, colors)
    if not show_legend:
        chart.legend = None
    if show_labels:
        add_data_labels(chart, number_format, position="outEnd")
    sheet.add_chart(chart, anchor)


def build_chart_sheets(workbook: Workbook, rows: list[dict]) -> None:
    clear_generated_sheets(workbook)
    data_sheet = workbook.create_sheet(CHART_DATA_SHEET_NAME)
    dashboard = workbook.create_sheet(CHART_SHEET_NAME, 0)
    data_sheet.sheet_state = "hidden"
    setup_dashboard_sheet(dashboard)

    chart_data = build_chart_data(rows)
    write_kpi_cards(dashboard, chart_data)

    sections = {
        "event": write_matrix(data_sheet, 1, 1, chart_data["event_rows"]),
        "balance": write_matrix(data_sheet, 1, 10, chart_data["balance_rows"]),
        "pmt_operation": write_matrix(data_sheet, 1, 14, chart_data["pmt_operation_rows"]),
        "monthly": write_matrix(data_sheet, 1, 18, chart_data["monthly_rows"]),
        "indexer": write_matrix(data_sheet, 1, 23, chart_data["indexer_rows"]),
        "returns": write_matrix(data_sheet, 1, 27, chart_data["return_rows"]),
        "kpis": write_matrix(data_sheet, 1, 33, chart_data["kpis"]),
        "event_pmt": write_matrix(data_sheet, 1, 37, chart_data["event_pmt_rows"]),
        "event_balance": write_matrix(data_sheet, 1, 41, chart_data["event_balance_rows"]),
    }

    for index, (name, bounds) in enumerate(sections.items(), start=1):
        style_table(data_sheet, bounds, f"tbl_{name}_{index}")

    for row in data_sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    for col in range(1, data_sheet.max_column + 1):
        data_sheet.column_dimensions[get_column_letter(col)].width = 18

    add_line_chart(dashboard, data_sheet, sections["event_pmt"], "A10", "PMT por evento com linha média (R$ mi)", 2, "1B998B", level_col=3)
    add_line_chart(dashboard, data_sheet, sections["event_balance"], "K10", "Saldo devedor projetado com nível médio (R$ mi)", 2, "F2994A", level_col=3)
    add_bar_chart(dashboard, data_sheet, sections["monthly"], "A32", "Composição mensal: juros x amortização (R$ mi)", 3, 4, ["4F83CC", "22C55E"], show_labels=False)
    add_bar_chart(dashboard, data_sheet, sections["balance"], "K32", "Ranking de saldo atual por emissão (R$ mi)", 2, 2, ["3B82F6"], chart_type="bar", show_legend=False)
    add_bar_chart(dashboard, data_sheet, sections["pmt_operation"], "A54", "PMT acumulado dos próximos 12 meses por operação (R$ mi)", 2, 2, ["0F766E"], chart_type="bar", show_legend=False)
    add_bar_chart(dashboard, data_sheet, sections["returns"], "K54", "TIR efetiva versus taxa contratada (% a.a.)", 2, 3, ["0B5CAD", "94A3B8"], number_format="0.0%")

    notes = [
        ("A29", "Linha pontilhada: média do período exibido. Rótulos em R$ mi para facilitar recorte direto em PPT."),
        ("K29", "Saldo em R$ mi; linha pontilhada mostra o nível médio da janela selecionada."),
        ("A51", "Barras mostram a abertura mensal dos próximos 12 meses entre juros e amortização."),
        ("K51", "Ranking em ordem decrescente, útil para destacar concentrações da carteira."),
        ("A75", "Janela móvel de 12 meses a partir da data de atualização diária."),
        ("K75", "Taxas anuais; comparar TIR efetiva contra taxa contratada por operação."),
    ]
    for cell_ref, text in notes:
        cell = dashboard[cell_ref]
        cell.value = text
        cell.font = Font(size=9, italic=True, color="607D95")

    dashboard.sheet_properties.tabColor = "0B5CAD"
    data_sheet.sheet_properties.tabColor = "94A3B8"


def ensure_workbook() -> Workbook:
    if TARGET_XLSX.exists():
        return load_workbook(TARGET_XLSX)
    workbook = Workbook()
    workbook.save(TARGET_XLSX)
    return workbook


def format_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    widths = {
        "A": 16,
        "B": 34,
        "C": 18,
        "D": 14,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 28,
        "K": 22,
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in ("F", "G", "H", "I"):
        for cell in sheet[column][1:]:
            cell.number_format = 'R$ #,##0.00'


def save_workbook_with_fallback(workbook: Workbook) -> tuple[Path, bool]:
    try:
        workbook.save(TARGET_XLSX)
        return TARGET_XLSX, False
    except PermissionError:
        fallback = TARGET_XLSX.with_name(
            f"{TARGET_XLSX.stem}_atualizado_{datetime.now().strftime('%Y%m%d_%H%M%S')}{TARGET_XLSX.suffix}"
        )
        workbook.save(fallback)
        return fallback, True


def write_rows(rows: list[dict]) -> tuple[Path, bool]:
    workbook = ensure_workbook()
    sheet = workbook["Plan1"] if "Plan1" in workbook.sheetnames else workbook.active
    sheet.title = "Plan1"
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    sheet.append(HEADERS)
    updated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for item in rows:
        sheet.append(
            [
                item["company_number"],
                item["company_name"],
                item["date_text"],
                item["operation"],
                item["emission"],
                item["payment"],
                item["interest"],
                item["amortization"],
                item["balance"],
                item["indexer"],
                updated_at,
            ]
        )
    format_sheet(sheet)
    build_chart_sheets(workbook, rows)
    return save_workbook_with_fallback(workbook)


def main() -> None:
    rows = collect_rows()
    saved_path, fallback = write_rows(rows)
    if fallback:
        print(f"PMT.xlsx estava aberta/bloqueada. Cópia atualizada gerada em {saved_path}")
    else:
        print(f"PMT.xlsx atualizada com {len(rows)} linhas futuras em {TARGET_XLSX}")


if __name__ == "__main__":
    main()
