# -*- coding: utf-8 -*-
"""
AXS Energia Unidade 05 - 2a emissao de debentures em 2 series
Calculo PMT padrao ANBIMA com projecao Focus/BCB.

Escopo desta versao
- Calcula a Primeira Serie e a Segunda Serie em um unico script.
- Mantem o padrao dos motores IPCA anteriores do projeto.
- Exporta arquivos separados por serie e um consolidado.

Premissas principais extraidas da escritura
- Data de emissao: 24/02/2026.
- Inicio da rentabilidade: 13/03/2026.
  Observacao: a escritura define a data como a primeira integralizacao. A data de
  13/03/2026 foi adotada com base nas informacoes publicas da emissao.
- Atualizacao monetaria: IPCA, aniversario no dia 15.
- Juros: base 252 dias uteis, com 3 incorporacoes iniciais em:
  15/08/2026, 15/02/2027 e 15/08/2027.
- Primeiro pagamento de juros e primeira amortizacao: 15/02/2028.
- Vencimento das 2 series: 15/02/2042.

Taxas adotadas
- Segunda Serie: 10,9659% a.a., conforme informacao publica da emissao.
- Primeira Serie: 9,4659% a.a., inferida a partir do mesmo bookbuilding da escritura:
  mesma NTN-B 2035 de referencia da emissao + spread de 1,60%, usando como base
  a taxa publica fechada da Segunda Serie (10,9659% = referencia + 3,10%).

Arquivos gerados
- controle_divida_axs05_v1_primeira_serie.csv
- controle_divida_axs05_v1_primeira_serie.xlsx
- controle_divida_axs05_v1_segunda_serie.csv
- controle_divida_axs05_v1_segunda_serie.xlsx
- controle_divida_axs05_v1_consolidado.csv
- controle_divida_axs05_v1_consolidado.xlsx
"""

from __future__ import annotations

import csv
import json
import ssl
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen
import importlib.util

getcontext().prec = 34

BASE_DIR = Path(__file__).resolve().parent

DATA_EMISSAO = date(2026, 2, 24)
DATA_INICIO_RENTABILIDADE = date(2026, 3, 13)
DATAS_INCORPORACAO_JUROS = [
    date(2026, 8, 15),
    date(2027, 2, 15),
    date(2027, 8, 15),
]
PU_INICIAL = Decimal("1000.00000000")
IPCA_PROJECAO_MENSAL_PADRAO = Decimal("0.0045")


CRONOGRAMA_RAW = [
    ("2028-02-15", "0.4630", "9.0096"),
    ("2028-08-15", "0.4859", "9.4559"),
    ("2029-02-15", "0.5076", "9.8909"),
    ("2029-08-15", "0.6156", "10.2985"),
    ("2030-02-15", "0.6586", "10.6116"),
    ("2030-08-15", "0.7131", "10.7959"),
    ("2031-02-15", "0.7380", "10.7683"),
    ("2031-08-15", "0.7877", "10.4367"),
    ("2032-02-15", "3.0793", "9.6539"),
    ("2032-08-15", "3.3881", "10.2411"),
    ("2033-02-15", "3.1106", "10.8548"),
    ("2033-08-15", "3.3917", "11.6107"),
    ("2034-02-15", "3.6912", "12.4267"),
    ("2034-08-15", "4.0729", "13.3002"),
    ("2035-02-15", "4.2206", "14.2036"),
    ("2035-08-15", "4.1950", "15.1819"),
    ("2036-02-15", "4.6055", "16.2855"),
    ("2036-08-15", "5.1307", "17.3485"),
    ("2037-02-15", "5.6956", "18.1893"),
    ("2037-08-15", "6.4193", "18.5042"),
    ("2038-02-15", "8.7899", "17.6919"),
    ("2038-08-15", "10.2516", "19.9136"),
    ("2039-02-15", "12.0569", "22.6821"),
    ("2039-08-15", "14.5759", "26.2517"),
    ("2040-02-15", "18.0228", "30.9962"),
    ("2040-08-15", "23.3761", "37.6572"),
    ("2041-02-15", "32.2442", "47.6373"),
    ("2041-08-15", "50.6040", "64.4761"),
    ("2042-02-15", "100.0000", "100.0000"),
]

# Fallback local apenas para nao interromper o calculo quando nao houver acesso externo.
IPCA_FALLBACK_VARIACAO_MENSAL = {
    "2024-02": Decimal("0.0083"), "2024-03": Decimal("0.0016"), "2024-04": Decimal("0.0038"),
    "2024-05": Decimal("0.0046"), "2024-06": Decimal("0.0021"), "2024-07": Decimal("0.0038"),
    "2024-08": Decimal("-0.0002"), "2024-09": Decimal("0.0044"), "2024-10": Decimal("0.0056"),
    "2024-11": Decimal("0.0039"), "2024-12": Decimal("0.0052"), "2025-01": Decimal("0.0016"),
    "2025-02": Decimal("0.0131"), "2025-03": Decimal("0.0056"), "2025-04": Decimal("0.0043"),
    "2025-05": Decimal("0.0026"), "2025-06": Decimal("0.0024"), "2025-07": Decimal("0.0026"),
    "2025-08": Decimal("-0.0011"), "2025-09": Decimal("0.0048"), "2025-10": Decimal("0.0009"),
    "2025-11": Decimal("0.0018"), "2025-12": Decimal("0.0033"), "2026-01": Decimal("0.0033"),
    "2026-02": Decimal("0.0070"), "2026-03": Decimal("0.0088"),
}


@dataclass(frozen=True)
class SerieConfig:
    id: str
    nome: str
    ticker: str
    isin: str
    quantidade: Decimal
    taxa_aa: Decimal
    cronograma: List[Tuple[date, Decimal]]
    observacao_taxa: str


def build_cronograma(indice: int) -> List[Tuple[date, Decimal]]:
    out: List[Tuple[date, Decimal]] = []
    for data_txt, primeira, segunda in CRONOGRAMA_RAW:
        valor = primeira if indice == 1 else segunda
        out.append((datetime.strptime(data_txt, "%Y-%m-%d").date(), Decimal(valor) / Decimal("100")))
    return out


SERIES = {
    "primeira": SerieConfig(
        id="primeira",
        nome="AXS 05 - Primeira Serie",
        ticker="AXSC12",
        isin="BRAXSCDBS015",
        quantidade=Decimal("64000"),
        taxa_aa=Decimal("0.094659"),
        cronograma=build_cronograma(1),
        observacao_taxa="Taxa anual inferida pelo mesmo bookbuilding da emissao: taxa publica da 2a serie menos 1,50 p.p. de spread.",
    ),
    "segunda": SerieConfig(
        id="segunda",
        nome="AXS 05 - Segunda Serie",
        ticker="AXSC22",
        isin="BRAXSCDBS023",
        quantidade=Decimal("22200"),
        taxa_aa=Decimal("0.109659"),
        cronograma=build_cronograma(2),
        observacao_taxa="Taxa anual adotada conforme informacao publica da emissao.",
    ),
}


def trunc_dec(x: Decimal, casas: int = 8) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_DOWN)


def round_dec(x: Decimal, casas: int = 2) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_HALF_UP)


def add_months(dt: date, months: int) -> date:
    y = dt.year + (dt.month - 1 + months) // 12
    m = (dt.month - 1 + months) % 12 + 1
    return date(y, m, 1)


def mes_str(dt: date) -> str:
    return f"{dt.year:04d}-{dt.month:02d}"


def aniversario(data_pagto: date) -> date:
    return date(data_pagto.year, data_pagto.month, 15)


def easter_date(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def feriados_nacionais(start_year: int = 2024, end_year: int = 2042) -> set[date]:
    fs: set[date] = set()
    for y in range(start_year, end_year + 1):
        pascoa = easter_date(y)
        fs.update({
            date(y, 1, 1),
            pascoa - timedelta(days=48),
            pascoa - timedelta(days=47),
            pascoa - timedelta(days=2),
            date(y, 4, 21),
            date(y, 5, 1),
            pascoa + timedelta(days=60),
            date(y, 9, 7),
            date(y, 10, 12),
            date(y, 11, 2),
            date(y, 11, 15),
            date(y, 11, 20),
            date(y, 12, 25),
        })
    return fs


FERIADOS = feriados_nacionais()


def eh_dia_util(dt: date) -> bool:
    return dt.weekday() < 5 and dt not in FERIADOS


def dias_uteis(inicio: date, fim: date) -> int:
    count = 0
    dt = inicio
    while dt < fim:
        if eh_dia_util(dt):
            count += 1
        dt += timedelta(days=1)
    return count


def periodo_final_sidra() -> str:
    max_aniv = max(aniversario(d) for config in SERIES.values() for d, _ in config.cronograma)
    fim = add_months(max_aniv, 1)
    return f"{fim.year:04d}{fim.month:02d}"


def indices_fallback() -> Dict[str, Decimal]:
    idx = Decimal("7000.0000000000000")
    indices: Dict[str, Decimal] = {}
    for m in sorted(IPCA_FALLBACK_VARIACAO_MENSAL):
        idx = idx * (Decimal("1") + IPCA_FALLBACK_VARIACAO_MENSAL[m])
        indices[m] = idx
    return indices


def obter_ipca_numero_indice_sidra() -> Tuple[Dict[str, Decimal], str]:
    fim = periodo_final_sidra()
    url = f"https://apisidra.ibge.gov.br/values/t/1737/n1/all/v/2266/p/202402-{fim}?formato=json"
    try:
        ctx = ssl.create_default_context()
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=12, context=ctx) as resp:
            dados = json.loads(resp.read().decode("utf-8"))
        indices: Dict[str, Decimal] = {}
        for item in dados[1:]:
            periodo = str(item.get("D3C", ""))
            valor = str(item.get("V", "")).replace(",", ".")
            if len(periodo) == 6 and valor not in ("", "...", "-"):
                indices[f"{periodo[:4]}-{periodo[4:]}"] = Decimal(valor)
        if not indices:
            raise RuntimeError("SIDRA retornou vazio ou layout inesperado")
        return indices, f"SIDRA/IBGE Tabela 1737 v/2266 | {url}"
    except Exception as exc:
        return indices_fallback(), f"FALLBACK local por variacao mensal IPCA; motivo: {exc}"


def decimal_ptbr(valor: object) -> Decimal | None:
    if valor is None:
        return None
    txt = str(valor).strip().replace("%", "").replace(" ", "")
    if not txt or txt in ("...", "-", "--"):
        return None
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        return None


def parse_mes_referencia(valor: object) -> str | None:
    if valor is None:
        return None
    txt = str(valor).strip()
    if not txt:
        return None
    if len(txt) >= 7 and txt[4] == "-" and txt[:4].isdigit() and txt[5:7].isdigit():
        return txt[:7]
    if "/" in txt:
        partes = txt.split("/")
        if len(partes) >= 2 and partes[0].isdigit() and partes[1].isdigit():
            mes = int(partes[0])
            ano = int(partes[1])
            if ano < 100:
                ano += 2000
            if 1 <= mes <= 12:
                return f"{ano:04d}-{mes:02d}"
    if len(txt) == 6 and txt.isdigit():
        return f"{txt[:4]}-{txt[4:]}"
    return None


def obter_json_url(url: str, timeout: int = 20) -> object:
    ctx = ssl.create_default_context()
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8"))


def build_odata_url(recurso: str, params: Dict[str, str], usar_parenteses: bool = True) -> str:
    base = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata"
    sufixo = "()" if usar_parenteses and not recurso.endswith(")") else ""
    return f"{base}/{recurso}{sufixo}?{urlencode(params, quote_via=quote)}"


def obter_focus_odata_mensal() -> Tuple[Dict[str, Decimal], str]:
    recursos = ["ExpectativaMercadoMensais", "ExpectativasMercadoMensais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                url = build_odata_url(recurso, params, usar_parenteses=usar_parenteses)
                dados = obter_json_url(url, timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[str, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    mes = parse_mes_referencia(item.get("DataReferencia"))
                    med = decimal_ptbr(item.get("Mediana"))
                    if not mes or med is None:
                        continue
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if mes not in out or data_pub > out[mes][0]:
                        out[mes] = (data_pub, taxa)
                if out:
                    nome = recurso + ("()" if usar_parenteses else "")
                    return {m: v[1] for m, v in out.items()}, f"Focus/BCB mensal via OData {nome}"
                erros.append(f"{recurso}: sem registros")
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus/BCB mensal OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_odata_anual() -> Tuple[Dict[int, Decimal], str]:
    recursos = ["ExpectativasMercadoAnuais", "ExpectativaMercadoAnuais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                url = build_odata_url(recurso, params, usar_parenteses=usar_parenteses)
                dados = obter_json_url(url, timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[int, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    ref = str(item.get("DataReferencia", "")).strip()
                    if not ref.isdigit():
                        continue
                    ano = int(ref)
                    med = decimal_ptbr(item.get("Mediana"))
                    if med is None:
                        continue
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if ano not in out or data_pub > out[ano][0]:
                        out[ano] = (data_pub, taxa)
                if out:
                    nome = recurso + ("()" if usar_parenteses else "")
                    return {a: v[1] for a, v in out.items()}, f"Focus/BCB anual via OData {nome}"
                erros.append(f"{recurso}: sem registros")
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus/BCB anual OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_python_bcb() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    try:
        if importlib.util.find_spec("bcb") is None:
            return {}, {}, "python-bcb nao instalado"
        from bcb import Expectativas  # type: ignore
        em = Expectativas()
    except Exception as exc:
        return {}, {}, f"python-bcb indisponivel: {exc}"

    mensagens: List[str] = []
    mensal: Dict[str, Decimal] = {}
    anual: Dict[int, Decimal] = {}

    try:
        ep = em.get_endpoint("ExpectativaMercadoMensais")
        df = (
            ep.query()
            .filter(ep.Indicador == "IPCA")
            .select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana)
            .orderby(ep.Data.desc())
            .limit(20000)
            .collect()
        )
        temp: Dict[str, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            mes = parse_mes_referencia(item.get("DataReferencia"))
            med = decimal_ptbr(item.get("Mediana"))
            if not mes or med is None:
                continue
            taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
            data_pub = str(item.get("Data", ""))
            if mes not in temp or data_pub > temp[mes][0]:
                temp[mes] = (data_pub, taxa)
        mensal = {m: v[1] for m, v in temp.items()}
        mensagens.append(f"python-bcb mensal: {len(mensal)} meses")
    except Exception as exc:
        mensagens.append(f"python-bcb mensal falhou: {exc}")

    try:
        ep = em.get_endpoint("ExpectativasMercadoAnuais")
        df = (
            ep.query()
            .filter(ep.Indicador == "IPCA")
            .select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana)
            .orderby(ep.Data.desc())
            .limit(20000)
            .collect()
        )
        temp2: Dict[int, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            ref = str(item.get("DataReferencia", "")).strip()
            med = decimal_ptbr(item.get("Mediana"))
            if not ref.isdigit() or med is None:
                continue
            ano = int(ref)
            taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
            data_pub = str(item.get("Data", ""))
            if ano not in temp2 or data_pub > temp2[ano][0]:
                temp2[ano] = (data_pub, taxa)
        anual = {a: v[1] for a, v in temp2.items()}
        mensagens.append(f"python-bcb anual: {len(anual)} anos")
    except Exception as exc:
        mensagens.append(f"python-bcb anual falhou: {exc}")

    return mensal, anual, " | ".join(mensagens)


def obter_focus_ipca() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    mensal, anual, fonte_py = obter_focus_python_bcb()
    fontes = [fonte_py]
    if not mensal:
        mensal, fonte_m = obter_focus_odata_mensal()
        fontes.append(fonte_m)
    if not anual:
        anual, fonte_a = obter_focus_odata_anual()
        fontes.append(fonte_a)
    return mensal, anual, " ; ".join(fontes)


def taxa_mensal_por_focus(
    mes: str,
    focus_mensal: Dict[str, Decimal],
    focus_anual: Dict[int, Decimal],
) -> Tuple[Decimal, str]:
    if mes in focus_mensal:
        return focus_mensal[mes], "Focus/BCB mensal"
    ano = int(mes[:4])
    if ano in focus_anual:
        taxa = (Decimal("1") + focus_anual[ano]) ** (Decimal("1") / Decimal("12")) - Decimal("1")
        return taxa, "Focus/BCB anual convertido para taxa mensal equivalente"
    return IPCA_PROJECAO_MENSAL_PADRAO, "fallback local padrao"


def preencher_indices_futuros(indices: Dict[str, Decimal]) -> Tuple[Dict[str, Decimal], Dict[str, str]]:
    out = dict(indices)
    fontes = {m: "SIDRA/IBGE ou fallback oficial" for m in out}
    focus_mensal, focus_anual, fonte_focus = obter_focus_ipca()

    ultimo_mes = max(out)
    ultimo_indice = out[ultimo_mes]
    mes_atual = add_months(date(int(ultimo_mes[:4]), int(ultimo_mes[5:7]), 1), 1)
    fim = mes_str(add_months(max(aniversario(d) for config in SERIES.values() for d, _ in config.cronograma), -1))

    while mes_str(mes_atual) <= fim:
        m = mes_str(mes_atual)
        if m in IPCA_FALLBACK_VARIACAO_MENSAL:
            taxa = IPCA_FALLBACK_VARIACAO_MENSAL[m]
            fonte_taxa = "fallback oficial/conhecido local"
        else:
            taxa, fonte_taxa = taxa_mensal_por_focus(m, focus_mensal, focus_anual)
        ultimo_indice = ultimo_indice * (Decimal("1") + taxa)
        out[m] = ultimo_indice
        fontes[m] = f"{fonte_taxa} | {fonte_focus}"
        mes_atual = add_months(mes_atual, 1)
    return out, fontes


def fator_ipca(indices: Dict[str, Decimal], data_aniv: date) -> Tuple[Decimal, str, str, Decimal, Decimal]:
    mes_nik = mes_str(add_months(data_aniv, -1))
    mes_nik_1 = mes_str(add_months(data_aniv, -2))
    if mes_nik not in indices or mes_nik_1 not in indices:
        raise RuntimeError(f"IPCA necessario nao disponivel: NIk={mes_nik}, NIk_1={mes_nik_1}.")
    ni_k = indices[mes_nik]
    ni_k_1 = indices[mes_nik_1]
    fator = trunc_dec(ni_k / ni_k_1, 8)
    return fator, mes_nik, mes_nik_1, ni_k, ni_k_1


def fator_ipca_prorata(
    indices: Dict[str, Decimal],
    data_aniv: date,
    inicio: date,
) -> Tuple[Decimal, str, str, Decimal, Decimal, int, int]:
    fator_cheio, mes_nik, mes_nik_1, ni_k, ni_k_1 = fator_ipca(indices, data_aniv)
    prev_m = add_months(data_aniv, -1)
    inicio_aniv = date(prev_m.year, prev_m.month, 15)
    dup = dias_uteis(inicio, data_aniv)
    dut = dias_uteis(inicio_aniv, data_aniv)
    bruto = Decimal(str(float(fator_cheio) ** (dup / dut)))
    return trunc_dec(bruto, 8), mes_nik, mes_nik_1, ni_k, ni_k_1, dup, dut


def fator_juros_252(du: int, taxa_aa: Decimal) -> Decimal:
    bruto = Decimal(str((1.0 + float(taxa_aa)) ** (du / 252.0)))
    return bruto.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_UP)


def proxima_data_aniversario(dt: date) -> date:
    if dt.day <= 15:
        return date(dt.year, dt.month, 15)
    prox = add_months(date(dt.year, dt.month, 1), 1)
    return date(prox.year, prox.month, 15)


def aplicar_ipca_ate(
    saldo_pu: Decimal,
    data_ipca_atual: date,
    data_aniv_alvo: date,
    indices: Dict[str, Decimal],
    fonte_mes: Dict[str, str],
) -> Tuple[Decimal, List[Dict[str, object]], date]:
    detalhes: List[Dict[str, object]] = []
    if data_aniv_alvo < data_ipca_atual:
        raise RuntimeError("data_aniv_alvo anterior a data_ipca_atual")

    prox = proxima_data_aniversario(data_ipca_atual)
    atual = data_ipca_atual
    saldo = trunc_dec(saldo_pu, 8)

    while prox <= data_aniv_alvo:
        if atual == prox:
            prox_m = add_months(prox, 1)
            prox = date(prox_m.year, prox_m.month, 15)
            continue

        if atual.day == 15 and atual < prox:
            fator_c, mes_nik, mes_nik_1, ni_k, ni_k_1 = fator_ipca(indices, prox)
            prev_m = add_months(prox, -1)
            inicio_aniv = date(prev_m.year, prev_m.month, 15)
            dup_ipca = dias_uteis(inicio_aniv, prox)
            dut_ipca = dup_ipca
        else:
            fator_c, mes_nik, mes_nik_1, ni_k, ni_k_1, dup_ipca, dut_ipca = fator_ipca_prorata(indices, prox, atual)

        saldo_ini = saldo
        saldo = trunc_dec(saldo * fator_c, 8)
        detalhes.append({
            "Data_Aniv_IPCA": prox.strftime("%d/%m/%Y"),
            "Mes_NIk": mes_nik,
            "Mes_NIk_1": mes_nik_1,
            "NIk": ni_k,
            "NIk_1": ni_k_1,
            "Fonte_NIk": fonte_mes.get(mes_nik, ""),
            "DUP_IPCA": dup_ipca,
            "DUT_IPCA": dut_ipca,
            "Fator_C_IPCA": fator_c,
            "PU_Antes_IPCA": saldo_ini,
            "PU_Apos_IPCA": saldo,
        })
        atual = prox
        prox_m = add_months(prox, 1)
        prox = date(prox_m.year, prox_m.month, 15)

    return saldo, detalhes, data_aniv_alvo


def calcular_fluxo_serie(
    config: SerieConfig,
    indices: Dict[str, Decimal],
    fonte_mes: Dict[str, str],
) -> List[Dict[str, object]]:
    saldo_pu = trunc_dec(PU_INICIAL, 8)
    data_ref_juros = DATA_INICIO_RENTABILIDADE
    data_ipca_atual = DATA_INICIO_RENTABILIDADE
    linhas: List[Dict[str, object]] = []

    for i, data_incorp in enumerate(DATAS_INCORPORACAO_JUROS, start=1):
        saldo_pu, detalhes_ipca, data_ipca_atual = aplicar_ipca_ate(
            saldo_pu,
            data_ipca_atual,
            data_incorp,
            indices,
            fonte_mes,
        )
        du_incorp = dias_uteis(data_ref_juros, data_incorp)
        fj_incorp = fator_juros_252(du_incorp, config.taxa_aa)
        pu_juros_incorp = trunc_dec(saldo_pu * (fj_incorp - Decimal("1")), 8)
        saldo_antes_incorp = saldo_pu
        saldo_pu = trunc_dec(saldo_pu + pu_juros_incorp, 8)
        linhas.append({
            "Instrumento": config.nome,
            "Ticker": config.ticker,
            "ISIN": config.isin,
            "Evento": f"Incorporacao Juros {i}",
            "Data_Ref": data_incorp.strftime("%d/%m/%Y"),
            "Data_Pgto": "",
            "DU_Juros": du_incorp,
            "Qtde_IPCA_Aplicados": len(detalhes_ipca),
            "Ultimo_Mes_NIk": detalhes_ipca[-1]["Mes_NIk"] if detalhes_ipca else "",
            "Ultimo_Mes_NIk_1": detalhes_ipca[-1]["Mes_NIk_1"] if detalhes_ipca else "",
            "Ultimo_Fator_C_IPCA": detalhes_ipca[-1]["Fator_C_IPCA"] if detalhes_ipca else "",
            "Fator_Juros": fj_incorp,
            "Taxa_AA": config.taxa_aa,
            "TAI_Amort": Decimal("0"),
            "PU_VNa_Ini": trunc_dec(PU_INICIAL, 8) if i == 1 else linhas[-1]["PU_VNa_Fim"],
            "PU_VNa_Atualizado": saldo_antes_incorp,
            "PU_Juros": pu_juros_incorp,
            "PU_Juros_Incorporado": pu_juros_incorp,
            "PU_Amort": Decimal("0"),
            "PU_Total_Pago": Decimal("0"),
            "PU_VNa_Fim": saldo_pu,
            "Juros_R$": Decimal("0.00"),
            "Amort_R$": Decimal("0.00"),
            "PMT_Total": Decimal("0.00"),
            "Saldo_Devedor_R$": round_dec(saldo_pu * config.quantidade, 2),
        })
        data_ref_juros = data_incorp

    for data_pagto, perc_amort in config.cronograma:
        data_aniv = aniversario(data_pagto)
        pu_vna_ini = trunc_dec(saldo_pu, 8)
        saldo_pu, detalhes_ipca, data_ipca_atual = aplicar_ipca_ate(
            saldo_pu,
            data_ipca_atual,
            data_aniv,
            indices,
            fonte_mes,
        )

        du = dias_uteis(data_ref_juros, data_pagto)
        fj = fator_juros_252(du, config.taxa_aa)
        pu_vna_atualizado = trunc_dec(saldo_pu, 8)
        pu_juros = trunc_dec(pu_vna_atualizado * (fj - Decimal("1")), 8)
        pu_amort = trunc_dec(pu_vna_atualizado * perc_amort, 8)
        pu_vna_fim = trunc_dec(pu_vna_atualizado - pu_amort, 8)

        juros_rs = round_dec(pu_juros * config.quantidade, 2)
        amort_rs = round_dec(pu_amort * config.quantidade, 2)
        pmt_rs = round_dec(juros_rs + amort_rs, 2)
        saldo_rs = round_dec(pu_vna_fim * config.quantidade, 2)

        linhas.append({
            "Instrumento": config.nome,
            "Ticker": config.ticker,
            "ISIN": config.isin,
            "Evento": "Pagamento",
            "Data_Ref": data_aniv.strftime("%d/%m/%Y"),
            "Data_Pgto": data_pagto.strftime("%d/%m/%Y"),
            "DU_Juros": du,
            "Qtde_IPCA_Aplicados": len(detalhes_ipca),
            "Ultimo_Mes_NIk": detalhes_ipca[-1]["Mes_NIk"] if detalhes_ipca else "",
            "Ultimo_Mes_NIk_1": detalhes_ipca[-1]["Mes_NIk_1"] if detalhes_ipca else "",
            "Ultimo_Fator_C_IPCA": detalhes_ipca[-1]["Fator_C_IPCA"] if detalhes_ipca else "",
            "Fator_Juros": fj,
            "Taxa_AA": config.taxa_aa,
            "TAI_Amort": perc_amort,
            "PU_VNa_Ini": pu_vna_ini,
            "PU_VNa_Atualizado": pu_vna_atualizado,
            "PU_Juros": pu_juros,
            "PU_Juros_Incorporado": Decimal("0"),
            "PU_Amort": pu_amort,
            "PU_Total_Pago": trunc_dec(pu_juros + pu_amort, 8),
            "PU_VNa_Fim": pu_vna_fim,
            "Juros_R$": juros_rs,
            "Amort_R$": amort_rs,
            "PMT_Total": pmt_rs,
            "Saldo_Devedor_R$": saldo_rs,
        })

        saldo_pu = pu_vna_fim
        data_ref_juros = data_pagto

    return linhas


def ordenar_por_data(linhas: List[Dict[str, object]]) -> List[Dict[str, object]]:
    def chave(row: Dict[str, object]) -> Tuple[date, str, str]:
        data_txt = str(row.get("Data_Pgto") or row.get("Data_Ref") or "")
        data_val = datetime.strptime(data_txt, "%d/%m/%Y").date()
        return data_val, str(row.get("Ticker", "")), str(row.get("Evento", ""))
    return sorted(linhas, key=chave)


def calcular_fluxos_series() -> Tuple[Dict[str, List[Dict[str, object]]], str]:
    indices, fonte = obter_ipca_numero_indice_sidra()
    indices, fonte_mes = preencher_indices_futuros(indices)
    out: Dict[str, List[Dict[str, object]]] = {}
    for serie_id, config in SERIES.items():
        out[serie_id] = calcular_fluxo_serie(config, indices, fonte_mes)
    return out, fonte


def calcular_fluxo(serie: str | None = None) -> Tuple[List[Dict[str, object]], str]:
    fluxos, fonte = calcular_fluxos_series()
    if serie:
        chave = serie.strip().lower()
        if chave not in fluxos:
            raise KeyError(f"Serie invalida: {serie}")
        return fluxos[chave], fonte
    consolidado = ordenar_por_data([linha for linhas in fluxos.values() for linha in linhas])
    return consolidado, fonte


def salvar_csv(linhas: List[Dict[str, object]], caminho: Path) -> None:
    if not linhas:
        raise RuntimeError("Nenhuma linha calculada.")
    campos = list(linhas[0].keys())
    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        w.writeheader()
        for row in linhas:
            out: Dict[str, object] = {}
            for k, v in row.items():
                out[k] = str(v).replace(".", ",") if isinstance(v, Decimal) else v
            w.writerow(out)


def _decimal_to_float(v: object) -> object:
    if isinstance(v, Decimal):
        return float(v)
    return v


def salvar_xlsx(linhas: List[Dict[str, object]], caminho: Path, titulo: str, fonte_ipca: str, observacao: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception as exc:
        raise RuntimeError("Para gerar XLSX, instale openpyxl: pip install openpyxl") from exc

    if not linhas:
        raise RuntimeError("Nenhuma linha calculada.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle_Divida"
    ws.freeze_panes = "A2"

    campos = list(linhas[0].keys())
    ws.append(campos)
    for row in linhas:
        ws.append([_decimal_to_float(row.get(c, "")) for c in campos])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    money_cols = {"Juros_R$", "Amort_R$", "PMT_Total", "Saldo_Devedor_R$"}
    pu_cols = {"PU_VNa_Ini", "PU_VNa_Atualizado", "PU_Juros", "PU_Juros_Incorporado", "PU_Amort", "PU_Total_Pago", "PU_VNa_Fim", "Ultimo_Fator_C_IPCA"}
    factor_cols = {"Fator_Juros", "Taxa_AA"}
    pct_cols = {"TAI_Amort"}

    for idx, nome in enumerate(campos, start=1):
        col = get_column_letter(idx)
        if nome in money_cols:
            number_format = "#,##0.00"
        elif nome in pu_cols:
            number_format = "0.00000000"
        elif nome in factor_cols:
            number_format = "0.000000000"
        elif nome in pct_cols:
            number_format = "0.0000%"
        elif nome in {"DU_Juros", "Qtde_IPCA_Aplicados"}:
            number_format = "0"
        else:
            number_format = None
        if number_format:
            for cell in ws[col][1:]:
                cell.number_format = number_format

    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.row_dimensions[1].height = 28

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="TabelaControleAXS05", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    for cell in ws[1]:
        if cell.value == "PU_Juros_Incorporado":
            cell.comment = Comment("Juros incorporados ao saldo nas tres datas de carencia previstas na escritura.", "Codex")
        if cell.value == "Taxa_AA":
            cell.comment = Comment("Taxa anual efetiva utilizada no fator de juros base 252.", "Codex")

    ws2 = wb.create_sheet("Parametros")
    params = [
        ["Campo", "Valor"],
        ["Titulo", titulo],
        ["Data de Emissao", DATA_EMISSAO.strftime("%d/%m/%Y")],
        ["Data de Inicio da Rentabilidade", DATA_INICIO_RENTABILIDADE.strftime("%d/%m/%Y")],
        ["Datas de Incorporacao dos Juros", ", ".join(d.strftime("%d/%m/%Y") for d in DATAS_INCORPORACAO_JUROS)],
        ["PU Inicial", float(PU_INICIAL)],
        ["Fonte IPCA", fonte_ipca],
        ["Observacao", observacao],
    ]
    for r in params:
        ws2.append(r)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 95
    ws2["B6"].number_format = "#,##0.00"

    wb.save(caminho)


def imprimir_linha(data_txt: str, linhas: List[Dict[str, object]], ticker: str | None = None) -> None:
    linha = next(
        (
            x for x in linhas
            if x["Data_Pgto"] == data_txt and (ticker is None or x.get("Ticker") == ticker)
        ),
        None,
    )
    if not linha:
        print(f"\nLinha {data_txt}: nao calculada.")
        return
    print(f"\nLinha {data_txt} - {linha.get('Instrumento')}:")
    campos = [
        "Ticker", "Evento", "Data_Ref", "Data_Pgto", "DU_Juros", "Qtde_IPCA_Aplicados",
        "Ultimo_Mes_NIk", "Ultimo_Mes_NIk_1", "Ultimo_Fator_C_IPCA",
        "Fator_Juros", "Taxa_AA", "PU_VNa_Atualizado", "PU_Juros", "PU_Amort",
        "Juros_R$", "Amort_R$", "PMT_Total", "Saldo_Devedor_R$",
    ]
    for k in campos:
        print(f"{k}: {linha.get(k, '')}")


def main() -> None:
    fluxos, fonte = calcular_fluxos_series()
    consolidado = ordenar_por_data([linha for linhas in fluxos.values() for linha in linhas])

    arquivos = {
        "primeira": "controle_divida_axs05_v1_primeira_serie",
        "segunda": "controle_divida_axs05_v1_segunda_serie",
        "consolidado": "controle_divida_axs05_v1_consolidado",
    }

    for serie_id in ("primeira", "segunda"):
        config = SERIES[serie_id]
        linhas = fluxos[serie_id]
        base_nome = arquivos[serie_id]
        salvar_csv(linhas, BASE_DIR / f"{base_nome}.csv")
        try:
            salvar_xlsx(
                linhas,
                BASE_DIR / f"{base_nome}.xlsx",
                config.nome,
                fonte,
                config.observacao_taxa,
            )
            print(f"Arquivo XLSX gerado: {base_nome}.xlsx")
        except Exception as exc:
            print(f"Nao foi possivel gerar XLSX de {config.nome}: {exc}")
        print(f"Arquivo CSV gerado: {base_nome}.csv")

    salvar_csv(consolidado, BASE_DIR / f"{arquivos['consolidado']}.csv")
    try:
        salvar_xlsx(
            consolidado,
            BASE_DIR / f"{arquivos['consolidado']}.xlsx",
            "AXS 05 - Consolidado das 2 series",
            fonte,
            "Consolidado das duas series da 2a emissao da AXS 05.",
        )
        print(f"Arquivo XLSX gerado: {arquivos['consolidado']}.xlsx")
    except Exception as exc:
        print(f"Nao foi possivel gerar XLSX consolidado: {exc}")
    print(f"Arquivo CSV gerado: {arquivos['consolidado']}.csv")

    print("Fonte IPCA oficial/historica:", fonte)
    print("Premissa taxa 1a serie:", SERIES["primeira"].observacao_taxa)
    imprimir_linha("15/02/2028", fluxos["primeira"], "AXSC12")
    imprimir_linha("15/02/2028", fluxos["segunda"], "AXSC22")
    imprimir_linha("15/02/2042", fluxos["segunda"], "AXSC22")


if __name__ == "__main__":
    main()
